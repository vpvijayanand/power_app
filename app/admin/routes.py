from flask import render_template, redirect, url_for, flash, request, jsonify, current_app
from app.auth.decorators import login_required, admin_required, get_current_user
from app.models import User, Instrument, OptionChainData, IndexData, NiftyPattern, db
from sqlalchemy import func
from app.admin import admin_bp
from datetime import datetime, timedelta
import io
import pandas as pd
from flask import send_file


@admin_bp.route('/')
@login_required
@admin_required
def dashboard():
    """Admin dashboard"""
    users = User.query.all()
    
    # Calculate statistics
    total_users = len(users)
    active_users = sum(1 for u in users if u.is_active)
    total_balance = sum(float(u.kite_account_balance) for u in users)
    avg_growth = sum(float(u.account_growth_percentage) for u in users) / total_users if total_users > 0 else 0
    
    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'total_balance': total_balance,
        'avg_growth': avg_growth
    }
    
    return render_template('admin/dashboard.html', 
                         users=users, 
                         stats=stats,
                         current_user=get_current_user())


@admin_bp.route('/toggle-user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_user(user_id):
    """Activate or deactivate a user"""
    user = db.session.get(User, user_id)
    
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Prevent admin from deactivating themselves
    current_user = get_current_user()
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'warning')
        return redirect(url_for('admin.dashboard'))
    
    # Toggle status
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'activated' if user.is_active else 'deactivated'
    flash(f'User {user.name} has been {status}.', 'success')
    
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/edit-user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    """Edit user details including Kite credentials"""
    user = db.session.get(User, user_id)
    
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    if request.method == 'POST':
        # Update user details
        user.name = request.form.get('name')
        user.email = request.form.get('email')
        user.mobile = request.form.get('mobile')
        user.trade_mode = request.form.get('trade_mode')
        user.lot_size = int(request.form.get('lot_size'))
        
        # Update Kite credentials if provided
        api_key = request.form.get('api_key')
        api_secret = request.form.get('api_secret')
        
        if api_key:
            user.api_key = api_key
        if api_secret:
            user.api_secret = api_secret
        
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash(f'User {user.name} updated successfully!', 'success')
        return redirect(url_for('admin.dashboard'))
    
@admin_bp.route('/instruments')
@login_required
@admin_required
def instruments():
    """View and export instruments data"""
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    exchange = request.args.get('exchange', '')
    instrument_type = request.args.get('type', '')
    segment = request.args.get('segment', '')
    expiry = request.args.get('expiry', '')
    action = request.args.get('action', '')
    
    # Build query
    query = Instrument.query
    
    # Apply filters
    if search:
        search_term = f"%{search}%"
        query = query.filter(db.or_(
            Instrument.tradingsymbol.ilike(search_term),
            Instrument.name.ilike(search_term),
            db.cast(Instrument.instrument_token, db.String).ilike(search_term)
        ))
    
    if exchange:
        query = query.filter(Instrument.exchange == exchange)
        
    if instrument_type:
        query = query.filter(Instrument.instrument_type == instrument_type)
        
    if segment:
        query = query.filter(Instrument.segment == segment)
        
    if expiry:
        query = query.filter(Instrument.expiry == expiry)
        
    # Get unique values for filters
    exchanges = db.session.query(Instrument.exchange).distinct().order_by(Instrument.exchange).all()
    types = db.session.query(Instrument.instrument_type).distinct().order_by(Instrument.instrument_type).all()
    segments = db.session.query(Instrument.segment).distinct().order_by(Instrument.segment).all()
    # Fetch future expiries for filter
    today = datetime.now().date()
    expiries = db.session.query(Instrument.expiry).filter(Instrument.expiry >= today).distinct().order_by(Instrument.expiry).all()
    
    exchanges = [e[0] for e in exchanges if e[0]]
    types = [t[0] for t in types if t[0]]
    segments = [s[0] for s in segments if s[0]]
    expiries = [e[0] for e in expiries if e[0]]
    
    # Handle Export
    if action == 'export':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Fetch all matching records (no pagination)
            instruments_data = query.all()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Instruments"
            
            headers = [
                'Token', 'Exchange Token', 'Symbol', 'Name', 'Last Price', 
                'Exchange', 'Type', 'Segment', 'Expiry', 'Strike', 
                'Tick Size', 'Lot Size', 'Fetch Date'
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row, inst in enumerate(instruments_data, 2):
                ws.cell(row=row, column=1, value=inst.instrument_token)
                ws.cell(row=row, column=2, value=inst.exchange_token)
                ws.cell(row=row, column=3, value=inst.tradingsymbol)
                ws.cell(row=row, column=4, value=inst.name)
                ws.cell(row=row, column=5, value=float(inst.last_price) if inst.last_price else 0)
                ws.cell(row=row, column=6, value=inst.exchange)
                ws.cell(row=row, column=7, value=inst.instrument_type)
                ws.cell(row=row, column=8, value=inst.segment)
                ws.cell(row=row, column=9, value=inst.expiry)
                ws.cell(row=row, column=10, value=float(inst.strike) if inst.strike else 0)
                ws.cell(row=row, column=11, value=float(inst.tick_size) if inst.tick_size else 0)
                ws.cell(row=row, column=12, value=inst.lot_size)
                ws.cell(row=row, column=13, value=inst.fetch_date)
            
            # Adjust widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'instruments_export_{timestamp}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            flash('Excel export requires openpyxl package.', 'danger')
            return redirect(url_for('admin.instruments'))
        except Exception as e:
            flash(f'Error exporting data: {str(e)}', 'danger')
            return redirect(url_for('admin.instruments'))

    # Pagination
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    instruments = pagination.items
    
    return render_template(
        'admin/instruments.html',
        instruments=instruments,
        pagination=pagination,
        exchanges=exchanges,
        types=types,
        segments=segments,
        expiries=expiries,
        filters={
            'search': search,
            'exchange': exchange,
            'type': instrument_type,
            'segment': segment,
            'expiry': expiry,
            'per_page': per_page
        },
        current_user=get_current_user()
    )


@admin_bp.route('/option-chain')
@login_required
@admin_required
def option_chain():
    """View and export option chain data"""
    # Get filter parameters
    underlying = request.args.get('underlying', 'NIFTY') # Default to NIFTY
    expiry = request.args.get('expiry', '')
    strike = request.args.get('strike', '')
    action = request.args.get('action', '')
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        query_date = datetime.now().date()
        date_str = query_date.strftime('%Y-%m-%d')
    
    # Get current spot
    current_spot = 0
    latest_ts = db.session.query(func.max(IndexData.timestamp)).filter(
        IndexData.symbol == underlying,
        func.date(IndexData.timestamp) == query_date
    ).scalar()
    
    if latest_ts:
        indx = db.session.query(IndexData).filter(
            IndexData.symbol == underlying,
            IndexData.timestamp == latest_ts
        ).first()
        if indx:
            current_spot = indx.close

    # Base query
    query = OptionChainData.query
    
    # Apply filters
    if underlying:
        query = query.filter(OptionChainData.underlying == underlying)
        
    # Get unique values for dropdowns
    underlyings = db.session.query(OptionChainData.underlying).distinct().order_by(OptionChainData.underlying).all()
    expiries = db.session.query(OptionChainData.expiry_date).filter(
        OptionChainData.expiry_date >= datetime.now().date()
    ).distinct().order_by(OptionChainData.expiry_date).all()
    
    underlyings = [u[0] for u in underlyings if u[0]]
    expiries = [e[0] for e in expiries if e[0]]
    
    # Default expiry if not selected and available
    if not expiry and expiries:
        expiry = str(expiries[0])
    
    # Apply expiry filter
    if expiry:
        query = query.filter(OptionChainData.expiry_date == expiry)

    # Apply DATE filter
    query = query.filter(func.date(OptionChainData.timestamp) == query_date)
        
    if strike:
        try:
            strike_val = float(strike)
            query = query.filter(OptionChainData.strike_price == strike_val)
        except ValueError:
            pass

    # Sort by Strike Price
    query = query.order_by(OptionChainData.strike_price.asc())

    # Handle Export
    if action == 'export':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            data = query.all()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{underlying} Options"
            
            # Headers
            headers = [
                'Time', 'Expiry', 'Strike', 
                'CE OI', 'CE Vol', 'CE LTP', 'CE Chg%', 'CE IV',
                'PE OI', 'PE Vol', 'PE LTP', 'PE Chg%', 'PE IV'
            ]
            
            # Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.timestamp)
                ws.cell(row=row, column=2, value=item.expiry_date)
                ws.cell(row=row, column=3, value=item.strike_price)
                
                # CE
                ws.cell(row=row, column=4, value=item.ce_oi)
                ws.cell(row=row, column=5, value=item.ce_volume)
                ws.cell(row=row, column=6, value=item.ce_ltp)
                ws.cell(row=row, column=7, value=item.ce_change_percent)
                ws.cell(row=row, column=8, value=item.ce_iv)
                
                # PE
                ws.cell(row=row, column=9, value=item.pe_oi)
                ws.cell(row=row, column=10, value=item.pe_volume)
                ws.cell(row=row, column=11, value=item.pe_ltp)
                ws.cell(row=row, column=12, value=item.pe_change_percent)
                ws.cell(row=row, column=13, value=item.pe_iv)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'option_chain_{underlying}_{timestamp}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            flash(f'Export error: {str(e)}', 'danger')
            return redirect(url_for('admin.option_chain'))

    # Fetch data for view
    option_data = query.all()
    
    return render_template(
        'admin/option_chain.html',
        option_data=option_data,
        underlyings=underlyings,
        expiries=expiries,
        filters={
            'underlying': underlying,
            'expiry': expiry,
            'strike': strike,
            'date': date_str
        },
        current_spot=current_spot
    )


@admin_bp.route('/option-analysis')
@login_required
@admin_required
def option_analysis():
    """Visualize option chain data"""
    underlying = request.args.get('underlying', 'NIFTY')
    expiry = request.args.get('expiry', '')
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    strike_range = request.args.get('range', 6, type=int)
    
    try:
        query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        query_date = datetime.now().date()
        date_str = query_date.strftime('%Y-%m-%d')
    
    # Get dropdown values
    underlyings = db.session.query(OptionChainData.underlying).distinct().order_by(OptionChainData.underlying).all()
    expiries = db.session.query(OptionChainData.expiry_date).filter(
        OptionChainData.expiry_date >= datetime.now().date()
    ).distinct().order_by(OptionChainData.expiry_date).all()
    
    underlyings = [u[0] for u in underlyings if u[0]]
    expiries = [e[0] for e in expiries if e[0]]
    
    # Default expiry
    if not expiry and expiries:
        expiry = str(expiries[0])
        
    data = []
    current_spot = 0
    
    if expiry:
        # 1. Get the latest timestamp for this selection
        latest_ts = db.session.query(func.max(OptionChainData.timestamp)).filter(
            OptionChainData.underlying == underlying,
            OptionChainData.expiry_date == expiry,
            func.date(OptionChainData.timestamp) == query_date
        ).scalar()
        
        if latest_ts:
            # 2. Fetch Spot Price from IndexData
            index_price = db.session.query(IndexData.close).filter(
                IndexData.symbol == underlying,
                IndexData.timestamp == latest_ts
            ).scalar()
            
            current_spot = index_price if index_price else 0
            
            # 3. Fetch Data for this Timestamp
            query = OptionChainData.query.filter(
                OptionChainData.underlying == underlying,
                OptionChainData.expiry_date == expiry,
                OptionChainData.timestamp == latest_ts
            )
            
            all_data = query.order_by(OptionChainData.strike_price.asc()).all()
            
            if all_data:
                # 4. Filter by Range
                # If Spot is 0 (missing index data), use middle of chain? 
                # Let's try to infer spot from ATM strike if index missing?
                # Or just show all if spot missing.
                
                if current_spot == 0 and all_data:
                     # Approximate spot using average of strikes? No, meaningless.
                     # Just take median strike.
                     mid_idx = len(all_data) // 2
                     current_spot = all_data[mid_idx].strike_price
                
                # Find closest strike to spot
                closest_strike_data = min(all_data, key=lambda x: abs(x.strike_price - current_spot))
                closest_strike = closest_strike_data.strike_price
                
                # Filter
                # We need N strikes below and N strikes above closest
                # Since all_data is sorted, we can find index of closest
                closest_idx = all_data.index(closest_strike_data)
                
                start_idx = max(0, closest_idx - strike_range)
                end_idx = min(len(all_data), closest_idx + strike_range + 1)
                
                data = all_data[start_idx:end_idx]
        
    return render_template(
        'admin/option_analysis.html',
        data=data,
        underlyings=underlyings,
        expiries=expiries,
        filters={
            'underlying': underlying,
            'expiry': expiry,
            'date': date_str,
            'range': strike_range
        },
        current_spot=current_spot
    )


@admin_bp.route('/option-crossover')
@login_required
@admin_required
def option_crossover():
    """Cross Over Analysis with Multi-Axis Chart"""
    underlying = request.args.get('underlying', 'NIFTY')
    expiry = request.args.get('expiry', '')
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    # Defaults
    try:
        query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        query_date = datetime.now().date()
        date_str = query_date.strftime('%Y-%m-%d')

    # Market hours boundaries (9:15 AM to 3:30 PM)
    market_open = datetime.combine(query_date, datetime.strptime('09:15', '%H:%M').time())
    market_close = datetime.combine(query_date, datetime.strptime('15:30', '%H:%M').time())

    # Get Filters
    underlyings = db.session.query(OptionChainData.underlying).distinct().order_by(OptionChainData.underlying).all()
    expiries = db.session.query(OptionChainData.expiry_date).filter(
        OptionChainData.expiry_date >= datetime.now().date()
    ).distinct().order_by(OptionChainData.expiry_date).all()
    
    underlyings = [u[0] for u in underlyings if u[0]]
    expiries = [e[0] for e in expiries if e[0]]
    
    if not expiry and expiries:
        expiry = str(expiries[0])

    # Get current spot
    current_spot = 0
    latest_ts = db.session.query(func.max(IndexData.timestamp)).filter(
        IndexData.symbol == underlying,
        func.date(IndexData.timestamp) == query_date
    ).scalar()
    
    if latest_ts:
        indx = db.session.query(IndexData).filter(
            IndexData.symbol == underlying,
            IndexData.timestamp == latest_ts
        ).first()
        if indx:
            current_spot = indx.close

    # 1. Fetch Option Chain Data (Aggregated by Timestamp)
    # We need SUM(ce_oi), SUM(pe_oi), SUM(ce_change), SUM(pe_change) per timestamp
    # Filtering by Underlying, Expiry, Date, and Market Hours (9:15 - 15:30)
    
    option_data = []
    if expiry:
        option_query = db.session.query(
            OptionChainData.timestamp,
            func.sum(OptionChainData.ce_oi).label('ce_oi_total'),
            func.sum(OptionChainData.pe_oi).label('pe_oi_total'),
            func.sum(OptionChainData.ce_oi_change).label('ce_change_total'),
            func.sum(OptionChainData.pe_oi_change).label('pe_change_total')
        ).filter(
            OptionChainData.underlying == underlying,
            OptionChainData.expiry_date == expiry,
            func.date(OptionChainData.timestamp) == query_date,
            OptionChainData.timestamp >= market_open,
            OptionChainData.timestamp <= market_close
        ).group_by(OptionChainData.timestamp).order_by(OptionChainData.timestamp)
        
        option_data = option_query.all()
    
    # 2. Fetch Index Data (Price)
    # Filter by Symbol (matches underlying), Date, and Market Hours (9:15 - 15:30)
    index_query = IndexData.query.filter(
        IndexData.symbol == underlying,
        func.date(IndexData.timestamp) == query_date,
        IndexData.timestamp >= market_open,
        IndexData.timestamp <= market_close
    ).order_by(IndexData.timestamp).all()
    
    # 3. Merge and Calculate Cumulative Change
    # To fix "jagged" graphs on restart, we calculate Change = Current Total OI - Open Total OI
    # We find the 'Open Total OI' from the first record of the day.
    
    data_map = {}
    
    # Base OIs
    base_ce_oi = None
    base_pe_oi = None
    
    if option_data:
        # First record establishes the baseline (Open OI)
        first_record = option_data[0]
        base_ce_oi = first_record.ce_oi_total or 0
        base_pe_oi = first_record.pe_oi_total or 0
    
    for row in option_data:
        ts = row.timestamp.strftime('%H:%M')
        if ts not in data_map:
            data_map[ts] = {}
            
        current_ce_oi = row.ce_oi_total or 0
        current_pe_oi = row.pe_oi_total or 0
        
        # Calculate Change from Baseline
        # If base is None (shouldn't happen if loop runs), use 0
        ce_change = current_ce_oi - (base_ce_oi if base_ce_oi is not None else 0)
        pe_change = current_pe_oi - (base_pe_oi if base_pe_oi is not None else 0)
        
        data_map[ts].update({
            'ce_oi': current_ce_oi,
            'pe_oi': current_pe_oi,
            'ce_change': ce_change, 
            'pe_change': pe_change
        })
        
    # Index Data processing
    for row in index_query:
        ts = row.timestamp.strftime('%H:%M')
        if ts not in data_map:
            data_map[ts] = {}
        
        # Calculate change (Price Momentum)
        change = row.close - row.open 
        data_map[ts]['price_change'] = change
        data_map[ts]['price'] = row.close
        
    # Convert back to sorted list (only market hours 09:15 - 15:30)
    chart_data = []
    sorted_keys = sorted(data_map.keys())
    
    for ts in sorted_keys:
        # Extra safety: skip any timestamps outside market hours
        if ts < '09:15' or ts > '15:30':
            continue

        item = data_map[ts]
        # Only include if we have at least one valid data point
        if not item: 
            continue
            
        chart_data.append({
            'time': ts,
            'ce_oi': item.get('ce_oi') or 0,
            'pe_oi': item.get('pe_oi') or 0,
            'ce_change': item.get('ce_change') or 0,
            'pe_change': item.get('pe_change') or 0,
            'nifty_change': item.get('price_change') or 0
        })

    return render_template(
        'admin/option_crossover.html',
        chart_data=chart_data,
        underlyings=underlyings,
        expiries=expiries,
        filters={
            'underlying': underlying,
            'expiry': expiry,
            'date': date_str
        },
        current_spot=current_spot
    )

@admin_bp.route('/delete-user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Delete a user"""
    user = db.session.get(User, user_id)
    
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Prevent admin from deleting themselves
    current_user = get_current_user()
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User {user.name} deleted successfully.', 'success')
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/index-analysis')
@login_required
@admin_required
def index_analysis():
    """Index Analysis Chart with Resampling"""
    
    # Filters
    symbol = request.args.get('symbol', 'NIFTY')
    duration = request.args.get('duration', '1m') # 1m, 2m, 3m, 5m, 10m, 15m
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    # Date Filtering
    query = IndexData.query.filter(IndexData.symbol == symbol)
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M')
            query = query.filter(IndexData.timestamp >= start_date)
        except ValueError:
            pass # Ignore invalid format

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M')
            query = query.filter(IndexData.timestamp <= end_date)
        except ValueError:
            pass

    # Default to today 9:15 AM to 3:30 PM if not provided
    if not start_date_str:
        now = datetime.now()
        start_date = datetime(now.year, now.month, now.day, 9, 15)
        start_date_str = start_date.strftime('%Y-%m-%dT%H:%M')
        query = query.filter(IndexData.timestamp >= start_date)

    if not end_date_str:
        now = datetime.now()
        end_date = datetime(now.year, now.month, now.day, 15, 30)
        end_date_str = end_date.strftime('%Y-%m-%dT%H:%M')
        query = query.filter(IndexData.timestamp <= end_date)
    
    # Fetch Data
    data = query.order_by(IndexData.timestamp).all()
    
    latest_spot = 0
    adx_status = "Neutral"
    adx_value = 0
    
    if not data:
        chart_data = []
    else:
        # Get Latest Data for Header
        latest_entry = data[-1]
        latest_spot = latest_entry.close
        adx_value = latest_entry.adx if latest_entry.adx else 0
        
        # ADX Logic
        if adx_value < 20:
            adx_status = "Weak Trend"
        elif adx_value > 25:
             # Determine direction from MA?
             if latest_entry.close > latest_entry.ma_20:
                 adx_status = "Strong Uptrend"
             else:
                 adx_status = "Strong Downtrend"
        else:
            adx_status = "Sideways / Building"

        # Convert to Pandas DataFrame
        df = pd.DataFrame([{
            'timestamp': d.timestamp,
            'open': d.open,
            'high': d.high,
            'low': d.low,
            'close': d.close,
            'ma_20': d.ma_20,
            'ma_200': d.ma_200,
            'atr': d.atr,
            'adx': d.adx
        } for d in data])
        
        df.set_index('timestamp', inplace=True)
        
        # Resample logic if duration > 1m
        if duration != '1m':
            # Map duration string to Pandas offset alias? 
            # 1m -> '1min', 5m -> '5min'
            rule = duration.replace('m', 'min')
            
            # OHLC Aggregation
            ohlc_dict = {
                'open': 'first',
                'high': 'max',
                'low': 'min',
                'close': 'last',
                'ma_20': 'last',   # Simplified: Take last value of the bucket
                'ma_200': 'last',  # Simplified
                'atr': 'last',     # Simplified
                'adx': 'last'      # Simplified
            }
            
            df_resampled = df.resample(rule).agg(ohlc_dict).dropna()
            df = df_resampled

        # Reset index to get timestamp back as column
        df.reset_index(inplace=True)
        
        # Prepare for JSON
        chart_data = df.to_dict(orient='records')
        # Convert Timestamp objects to string or ISO format for JS
        # And round floats
        for row in chart_data:
            if isinstance(row['timestamp'], datetime):
                 row['timestamp'] = row['timestamp'].isoformat()
            
            for key in ['open', 'high', 'low', 'close', 'ma_20', 'ma_200', 'atr', 'adx']:
                if row.get(key) is not None:
                    row[key] = round(float(row[key]), 2)
    
    # --- OI Analysis Logic ---
    # 1. Get Current Total OI Change for today (Assuming today's expiry or next expiry? Usually 'Current Expiry')
    # We need to find the active expiry for the symbol.
    today = datetime.now().date()
    
    # Find nearest expiry
    nearest_expiry = db.session.query(func.min(OptionChainData.expiry_date)).filter(
        OptionChainData.underlying == symbol,
        OptionChainData.expiry_date >= today
    ).scalar()
    
    oi_stats = {
        'ce_total_change': 0,
        'pe_total_change': 0,
        'diff_percent': 0,
        'signal': 'Neutral', # Red (Down) or Green (Up)
        'trend_text': 'Waiting for data...'
    }

    if nearest_expiry:
        # Get latest timestamp for this expiry to ensure we sum the "current" snapshot
        latest_ts = db.session.query(func.max(OptionChainData.timestamp)).filter(
            OptionChainData.underlying == symbol,
            OptionChainData.expiry_date == nearest_expiry
        ).scalar()
        
        if latest_ts:
            # Sum OI and OI Change for this timestamp
            sums = db.session.query(
                func.sum(OptionChainData.ce_oi),
                func.sum(OptionChainData.pe_oi),
                func.sum(OptionChainData.ce_oi_change),
                func.sum(OptionChainData.pe_oi_change)
            ).filter(
                OptionChainData.underlying == symbol,
                OptionChainData.expiry_date == nearest_expiry,
                OptionChainData.timestamp == latest_ts
            ).first()
            
            # Full OI
            ce_oi = sums[0] if sums[0] else 0
            pe_oi = sums[1] if sums[1] else 0
            oi_stats['ce_oi'] = ce_oi
            oi_stats['pe_oi'] = pe_oi
            
            # Full OI Diff
            diff_oi = abs(ce_oi - pe_oi)
            div_oi = max(abs(ce_oi), abs(pe_oi)) if max(abs(ce_oi), abs(pe_oi)) > 0 else 1
            oi_stats['oi_diff_percent'] = round((diff_oi / div_oi) * 100, 2)
            oi_stats['oi_signal'] = 'Bullish' if pe_oi > ce_oi else 'Bearish' # PE > CE = Support = Bullish

            # OI Change
            ce_change = sums[2] if sums[2] else 0
            pe_change = sums[3] if sums[3] else 0
            oi_stats['ce_total_change'] = ce_change
            oi_stats['pe_total_change'] = pe_change
            
            # Change Diff
            diff_change = abs(ce_change - pe_change)
            div_change = max(abs(ce_change), abs(pe_change)) if max(abs(ce_change), abs(pe_change)) > 0 else 1
            oi_stats['diff_percent'] = round((diff_change / div_change) * 100, 2)
            
            if ce_change > pe_change:
                oi_stats['signal'] = 'Bearish' # Red
            elif pe_change > ce_change:
                oi_stats['signal'] = 'Bullish' # Green
            
            # --- Historical Trend (15m / 30m) ---
            # Logic: Compare Current OI Change Sum vs Sum 15 mins ago vs Sum 30 mins ago
            
            # Helper to get sum at time T
            def get_oi_sum_at(time_target):
                # Find closest timestamp <= time_target
                ts = db.session.query(func.max(OptionChainData.timestamp)).filter(
                    OptionChainData.underlying == symbol,
                    OptionChainData.expiry_date == nearest_expiry,
                    OptionChainData.timestamp <= time_target
                ).scalar()
                
                if not ts: return 0, 0
                
                s = db.session.query(
                    func.sum(OptionChainData.ce_oi_change),
                    func.sum(OptionChainData.pe_oi_change)
                ).filter(
                    OptionChainData.underlying == symbol,
                    OptionChainData.expiry_date == nearest_expiry,
                    OptionChainData.timestamp == ts
                ).first()
                return (s[0] or 0), (s[1] or 0)

            ce_15, pe_15 = get_oi_sum_at(latest_ts - timedelta(minutes=15))
            # ce_30, pe_30 = get_oi_sum_at(latest_ts - timedelta(minutes=30)) # Optional if needed
            
            # Determine Trend Direction (Simple logic comparing Now vs 15m ago)
            # Threshold for "Consistent" vs "Increasing/Reducing"
            threshold = 0.05 # 5% change? or absolute? Let's use simple logic.
            
            def get_trend(current, past):
                if past == 0: return "Increasing" if current > 0 else "Consistent"
                change = (current - past) / abs(past)
                if change > 0.02: return "Increasing"
                if change < -0.02: return "Reducing"
                return "Consistent"

            ce_trend = get_trend(ce_change, ce_15)
            pe_trend = get_trend(pe_change, pe_15)
            
            # User Rules
            msg = "Analysing..."
            # 1. CE & PE Constant
            if ce_trend == "Consistent" and pe_trend == "Consistent":
                msg = "Trend Continuous"
            # 2. CE Reducing, PE Increasing
            elif ce_trend == "Reducing" and pe_trend == "Increasing":
                msg = "Strong Upside Move" # PE increasing (Support up), CE reducing (Resist down) -> Bullish
            # 3. CE Increasing, PE Reducing
            elif ce_trend == "Increasing" and pe_trend == "Reducing":
                msg = "Strong Downside Move" # CE increasing (Resist up), PE reducing -> Bearish
            # 4. CE Constant, PE Increasing
            elif ce_trend == "Consistent" and pe_trend == "Increasing":
                msg = "PE is Consistent, sudden upside risk" # Wait, user said: "if PE is Consistant and CE is increasing show PE is consistant so could be a sudden upside move" 
                # Re-reading user request:
                # "if CE is consistnat and PE is increating show CE is consistand so could be a sudden downside move" -> Wait, logic?
                # Usually PE Increasing = Bullish/Support. CE Increasing = Bearish/Resist.
                # User's text: "if CE is consistnat and PE is increating show CE is consistand so could be a sudden downside move" 
                # -> If PE increases (Support builds), usually that's UPSIDE. User says DOWNSIDE? 
                # Let's FOLLOW USER TEXT STRICTLY.
            
            # Mapping User Rules EXACTLY:
            # "if CE is reducing and PE in increating show text as trend is moving towards upwards" -> OK
            # "if CE is increasing and PE is reducing show as trend moving towards downwards" -> OK
            # "if CE is consistnat and PE is increating show CE is consistand so could be a sudden downside move"
            # "if PE is Consistant and CE is increasing show PE is consistant so could be a sudden upside move"
            # "if PE is reducing and CE is increaing show stong downside move" (Duplicate of 3?)
            # "if Ce is reducing and PE is increating show a strong upside move" (Duplicate of 2?)
            
            if ce_trend == "Reducing" and pe_trend == "Increasing":
                 msg = "Trend Moving Towards Upwards"
            elif ce_trend == "Increasing" and pe_trend == "Reducing":
                 msg = "Trend Moving Towards Downwards"
            elif ce_trend == "Consistent" and pe_trend == "Increasing":
                 msg = "CE is Consistent, could be sudden DOWNSIDE move" # User rule
            elif pe_trend == "Consistent" and ce_trend == "Increasing":
                 msg = "PE is Consistent, could be sudden UPSIDE move" # User rule
            else:
                 msg = f"CE: {ce_trend}, PE: {pe_trend}"

            # Determine Color based on message content
            trend_color = "text-white"
            if "Upwards" in msg or "UPSIDE" in msg or "Upside" in msg:
                trend_color = "text-green"
            elif "Downwards" in msg or "DOWNSIDE" in msg or "Downside" in msg:
                trend_color = "text-red"
            elif "Continuous" in msg:
                trend_color = "text-yellow"
            
            oi_stats['trend_text'] = msg
            oi_stats['trend_color'] = trend_color

    # Get symbols for dropdown
    symbols = db.session.query(IndexData.symbol).distinct().all()
    symbols = [s[0] for s in symbols]
    
    return render_template(
        'admin/index_chart.html',
        chart_data=chart_data,
        symbols=symbols,
        latest_spot=latest_spot,
        adx_status=adx_status,
        adx_value=adx_value,
        oi_stats=oi_stats,
        filters={
            'symbol': symbol,
            'duration': duration,
            'start_date': start_date_str,
            'end_date': end_date_str
        }
    )
    """Delete a user"""
    user = db.session.get(User, user_id)
    
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Prevent admin from deleting themselves
    current_user = get_current_user()
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'warning')
        return redirect(url_for('admin.dashboard'))
    
    user_name = user.name
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User {user_name} deleted successfully.', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/export-users')
@login_required
@admin_required
def export_users():
    """Export all users to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Header style
        header_fill = PatternFill(start_color="00AAFF", end_color="00AAFF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        # Headers
        headers = ['S.No', 'Name', 'Email', 'Mobile', 'User Type', 'Trade Mode', 
                   'Lot Size', 'Account Balance', 'Growth %', 'Status', 
                   'API Key', 'Has Request Token', 'Created At', 'Last Balance Update']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        users = User.query.all()
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=row-1)  # Serial number
            ws.cell(row=row, column=2, value=user.name)
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.mobile or 'N/A')
            ws.cell(row=row, column=5, value=user.user_type)
            ws.cell(row=row, column=6, value=user.trade_mode)
            ws.cell(row=row, column=7, value=user.lot_size)
            ws.cell(row=row, column=8, value=float(user.kite_account_balance))
            ws.cell(row=row, column=9, value=float(user.account_growth_percentage))
            ws.cell(row=row, column=10, value='Active' if user.is_active else 'Inactive')
            ws.cell(row=row, column=11, value=user.api_key or 'Not Set')
            ws.cell(row=row, column=12, value='Yes' if user.request_token else 'No')
            ws.cell(row=row, column=13, value=user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else 'N/A')
            ws.cell(row=row, column=14, value=user.last_balance_update.strftime('%Y-%m-%d %H:%M') if user.last_balance_update else 'N/A')
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        filename = f'users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except ImportError:
        flash('Excel export requires openpyxl package. Please install it.', 'danger')
        return redirect(url_for('admin.dashboard'))
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'danger')
        return redirect(url_for('admin.dashboard'))


# ===================================================================
# Helpers shared by OI Chart routes
# ===================================================================
# ===================================================================
# REPLACE the two helpers + oi_chart() + oi_chart_data() in routes.py
# ===================================================================

def _oi_chart_nearest_expiry(underlying='NIFTY', query_date=None):
    """
    Return nearest expiry >= query_date from OptionChainData as plain string.
    Uses string comparison so SQLite text columns always match.
    """
    if query_date is None:
        query_date = datetime.now().date()

    date_str = query_date.strftime('%Y-%m-%d')

    raw = db.session.query(
        func.min(OptionChainData.expiry_date)
    ).filter(
        OptionChainData.underlying  == underlying,
        OptionChainData.expiry_date >= date_str        # string >= string
    ).scalar()

    return str(raw) if raw else None


def _oi_chart_spot(underlying, query_date):
    """Return latest close for underlying on query_date."""
    latest_ts = db.session.query(func.max(IndexData.timestamp)).filter(
        IndexData.symbol == underlying,
        func.date(IndexData.timestamp) == query_date
    ).scalar()
    if latest_ts:
        val = db.session.query(IndexData.close).filter(
            IndexData.symbol    == underlying,
            IndexData.timestamp == latest_ts
        ).scalar()
        return round(float(val), 2) if val else 0
    return 0


def _oi_build_strikes_100(spot):
    """
    Return 7 strikes exclusively in 100-point steps: 
    ATM-300, ATM-200, ATM-100, ATM, ATM+100, ATM+200, ATM+300
    """
    step = 100
    atm  = round(spot / step) * step if spot else 23100
    return [atm + (i * step) for i in range(-3, 4)]


# ===================================================================
# ROUTE: OI Chart page
# ===================================================================
@admin_bp.route('/oi-chart')
@login_required
@admin_required
def oi_chart():
    today      = datetime.now().date()
    date_str   = request.args.get('date', today.strftime('%Y-%m-%d'))
    start_time = request.args.get('start_time', '09:15')
    end_time   = request.args.get('end_time', '15:30')
    underlying = 'NIFTY'

    try:
        query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        query_date = today
        date_str   = query_date.strftime('%Y-%m-%d')

    expiry  = _oi_chart_nearest_expiry(underlying, query_date)
    spot    = _oi_chart_spot(underlying, query_date)
    strikes = _oi_build_strikes_100(spot)

    return render_template(
        'admin/oi_chart.html',
        underlying=underlying,
        expiry=expiry or '',
        spot=spot,
        strikes=strikes,
        filters={
            'date':       date_str,
            'start_time': start_time,
            'end_time':   end_time,
        }
    )


# ===================================================================
# API: OI Chart data (JSON for Chart.js)
# ===================================================================
@admin_bp.route('/oi-chart-data')
@login_required
@admin_required
def oi_chart_data():
    import traceback as _tb
    try:
        underlying     = request.args.get('underlying', 'NIFTY')
        date_str       = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        start_time_str = request.args.get('start_time', '09:15')
        end_time_str   = request.args.get('end_time', '15:30')

        try:
            query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            query_date = datetime.now().date()
            date_str   = query_date.strftime('%Y-%m-%d')

        try:
            market_open  = datetime.combine(query_date, datetime.strptime(start_time_str, '%H:%M').time())
            market_close = datetime.combine(query_date, datetime.strptime(end_time_str,   '%H:%M').time())
        except ValueError:
            market_open  = datetime.combine(query_date, datetime.strptime('09:15', '%H:%M').time())
            market_close = datetime.combine(query_date, datetime.strptime('15:30', '%H:%M').time())

        # ── 1. Nearest expiry for THIS date (only one, nearest) ───────────
        expiry_str = _oi_chart_nearest_expiry(underlying, query_date)
        current_app.logger.info(f"[OI-CHART] Nearest expiry = {expiry_str!r}")

        if not expiry_str:
            return jsonify({
                'error': 'No expiry found in OptionChainData for this date/underlying.',
                'timestamps': [], 'nifty': [], 'strikes': {}, 'strike_list': [],
                'spot': 0, 'expiry': '', '_debug': {}
            })

        # ── 2. Spot ───────────────────────────────────────────────────────
        spot = _oi_chart_spot(underlying, query_date)

        # ── 3. Strikes available in DB for THIS expiry & date ─────────────
        #    Filter by expiry_str so we only see strikes for nearest expiry.
        avail_rows = db.session.query(OptionChainData.strike_price).filter(
            OptionChainData.underlying  == underlying,
            OptionChainData.expiry_date == expiry_str,     # ← scoped to nearest expiry
            func.date(OptionChainData.timestamp) == query_date
        ).distinct().order_by(OptionChainData.strike_price).all()

        available = [float(r[0]) for r in avail_rows if r[0] is not None]
        current_app.logger.info(
            f"[OI-CHART] Available strikes for expiry {expiry_str}: "
            f"count={len(available)}  sample={available[:10]}"
        )

        # ── 4. Pick 7 strikes around ATM ──────────────────────────────────
        strikes = _oi_build_strikes_100(spot)
        current_app.logger.info(f"[OI-CHART] spot={spot}  chosen strikes={[int(s) for s in strikes]}")

        # ── 5. Diagnostic counts ──────────────────────────────────────────
        any_index = db.session.query(func.count(IndexData.id)).filter(
            IndexData.symbol == underlying,
            func.date(IndexData.timestamp) == query_date
        ).scalar() or 0

        any_oc_total = db.session.query(func.count(OptionChainData.id)).filter(
            OptionChainData.underlying == underlying,
            func.date(OptionChainData.timestamp) == query_date
        ).scalar() or 0

        # Rows for the chosen expiry only
        any_oc_expiry = db.session.query(func.count(OptionChainData.id)).filter(
            OptionChainData.underlying  == underlying,
            OptionChainData.expiry_date == expiry_str,
            func.date(OptionChainData.timestamp) == query_date
        ).scalar() or 0

        # Rows for chosen expiry + chosen strikes (pre-window)
        any_oc_strikes = db.session.query(func.count(OptionChainData.id)).filter(
            OptionChainData.underlying   == underlying,
            OptionChainData.expiry_date  == expiry_str,
            OptionChainData.strike_price.in_(strikes),
            func.date(OptionChainData.timestamp) == query_date
        ).scalar() or 0

        current_app.logger.info(
            f"[OI-CHART] IndexData rows={any_index} | "
            f"OC total={any_oc_total} | OC expiry={any_oc_expiry} | "
            f"OC expiry+strikes={any_oc_strikes}"
        )

        # ── 6. Fetch OI rows ──────────────────────────────────────────────
        raw_rows = db.session.query(
            OptionChainData.timestamp,
            OptionChainData.strike_price,
            OptionChainData.ce_oi,
            OptionChainData.pe_oi,
        ).filter(
            OptionChainData.underlying   == underlying,
            OptionChainData.expiry_date  == expiry_str,
            OptionChainData.strike_price.in_(strikes),
            func.date(OptionChainData.timestamp) == query_date
        ).order_by(OptionChainData.timestamp).all()

        current_app.logger.info(f"[OI-CHART] raw_rows before window filter = {len(raw_rows)}")

        # ── 7. Parse timestamps & window filter ───────────────────────────
        def _parse_ts(val):
            if isinstance(val, datetime):
                return val
            if not val:
                return None
            val_str = str(val).split('+')[0].split('.')[0].replace('T', ' ')
            try:
                return datetime.strptime(val_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    return datetime.strptime(val_str, '%Y-%m-%d %H:%M')
                except ValueError:
                    return None

        rows = []
        for r in raw_rows:
            ts_val = _parse_ts(r.timestamp)
            if ts_val and market_open <= ts_val <= market_close:
                rows.append({
                    'parsed_ts':    ts_val,
                    'strike_price': float(r.strike_price),
                    'ce_oi':        int(r.ce_oi or 0),
                    'pe_oi':        int(r.pe_oi or 0),
                })

        current_app.logger.info(f"[OI-CHART] OI rows inside window = {len(rows)}")

        # ── 8. Build OI-change series (baseline = first row per strike) ────
        strike_keys      = [str(int(s)) for s in strikes]
        strike_snapshots = {sk: {} for sk in strike_keys}
        baseline_oi      = {}

        for r in rows:
            sk     = str(int(r['strike_price']))
            ts_str = r['parsed_ts'].strftime('%H:%M')

            if sk not in baseline_oi:
                baseline_oi[sk] = {'ce': r['ce_oi'], 'pe': r['pe_oi']}

            if sk in strike_snapshots:
                strike_snapshots[sk][ts_str] = {
                    'ce_oi': r['ce_oi'] - baseline_oi[sk]['ce'],
                    'pe_oi': r['pe_oi'] - baseline_oi[sk]['pe'],
                }

        oi_timestamps = sorted(set(
            ts for sk in strike_snapshots for ts in strike_snapshots[sk]
        ))

        # ── 9. Fetch Nifty price series ───────────────────────────────────
        index_rows = db.session.query(
            IndexData.timestamp,
            IndexData.close,
        ).filter(
            IndexData.symbol == underlying,
            func.date(IndexData.timestamp) == query_date,
            IndexData.timestamp >= market_open,
            IndexData.timestamp <= market_close,
        ).order_by(IndexData.timestamp).all()

        nifty_map = {}
        for r in index_rows:
            ts_val = _parse_ts(r.timestamp)
            if ts_val:
                nifty_map[ts_val.strftime('%H:%M')] = round(float(r.close), 2)

        # ── 10. Merge timelines ───────────────────────────────────────────
        all_timestamps = sorted(set(oi_timestamps) | set(nifty_map.keys()))
        nifty_series   = [nifty_map.get(ts) for ts in all_timestamps]

        strikes_data = {}
        for sk in strike_keys:
            strikes_data[sk] = {
                'ce_oi': [strike_snapshots[sk].get(ts, {}).get('ce_oi') for ts in all_timestamps],
                'pe_oi': [strike_snapshots[sk].get(ts, {}).get('pe_oi') for ts in all_timestamps],
            }

        # Fetch sample of expiries and strikes for the JS debug dump
        debug_oc_exp = db.session.query(OptionChainData.expiry_date).filter(
            func.date(OptionChainData.timestamp) == query_date
        ).distinct().all()
        debug_expiries = [str(e[0]) for e in debug_oc_exp if e[0]]

        debug_oc_str = db.session.query(OptionChainData.strike_price).filter(
            func.date(OptionChainData.timestamp) == query_date,
            OptionChainData.expiry_date == expiry_str
        ).distinct().order_by(OptionChainData.strike_price).all()
        debug_strikes = [float(s[0]) for s in debug_oc_str[:20] if s[0] is not None]

        return jsonify({
            'timestamps':  all_timestamps,
            'nifty':       nifty_series,
            'strikes':     strikes_data,
            'strike_list': strike_keys,
            'spot':        spot,
            'expiry':      expiry_str,
            '_debug': {
                'any_index_rows':     any_index,
                'any_oc_rows':        any_oc_total,          # ← Matches frontend JS
                'oc_expiries':        debug_expiries,        # ← Matches frontend JS
                'oc_strikes_sample':  debug_strikes,         # ← Matches frontend JS
                'expiry_used':        expiry_str,
                'strikes_queried':    [int(s) for s in strikes],
                'oi_rows_fetched':    len(rows),             # ← Matches frontend JS 
                'nifty_rows_fetched': len(index_rows),       # ← Matches frontend JS
                'raw_rows_found':     len(raw_rows),         
                'window_open':        str(market_open),
                'window_close':       str(market_close),
            }
        })

    except Exception as exc:
        err_msg = str(exc)
        tb_str  = _tb.format_exc()
        current_app.logger.error(f"[OI-CHART] EXCEPTION: {err_msg}\n{tb_str}")
        return jsonify({'error': err_msg, 'traceback': tb_str}), 500


@admin_bp.route('/index-patterns')
@login_required
@admin_required
def index_patterns():
    """Visualize the most frequent chart patterns formed"""
    symbol = request.args.get('symbol', 'NIFTY')
    pattern_type = request.args.get('pattern_type', '')
    
    query = NiftyPattern.query
    if symbol:
        query = query.filter(NiftyPattern.symbol == symbol)
    if pattern_type:
        query = query.filter(NiftyPattern.pattern_type == pattern_type)
        
    # Get distinct pattern types for filter
    types_query = db.session.query(NiftyPattern.pattern_type).filter(NiftyPattern.symbol == symbol).distinct().all()
    types = [t[0] for t in types_query if t[0]]
    
    # Get top 12 patterns by similar_count
    patterns = query.order_by(NiftyPattern.similar_count.desc()).limit(12).all()
    
    return render_template(
        'admin/index_patterns.html',
        patterns=patterns,
        symbols=['NIFTY', 'BANKNIFTY'],
        types=types,
        filters={'symbol': symbol, 'pattern_type': pattern_type}
    )

@admin_bp.route('/live-pattern-match')
@login_required
@admin_required
def live_pattern_match():
    """Visualize how today's shape matches historical patterns"""
    symbol = request.args.get('symbol', 'NIFTY')
    
    # Date and Time handling
    date_str = request.args.get('date', '')
    try:
        from datetime import time
        query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        query_date = datetime.now().date()
        date_str = query_date.strftime('%Y-%m-%d')
        
    start_time_str = request.args.get('start_time', '09:15')
    end_time_str = request.args.get('end_time', '15:30')
    
    try:
        from datetime import time
        ts_start = datetime.combine(query_date, datetime.strptime(start_time_str, '%H:%M').time())
        ts_end = datetime.combine(query_date, datetime.strptime(end_time_str, '%H:%M').time())
    except ValueError:
        from datetime import time
        ts_start = datetime.combine(query_date, time(9, 15))
        ts_end = datetime.combine(query_date, time(15, 30))
        start_time_str = '09:15'
        end_time_str = '15:30'
        
    threshold = request.args.get('threshold', 80.0, type=float)
    
    # Load engine visually from scripts without strict import errors locally
    import importlib.util
    import os
    engine_path = os.path.join(current_app.root_path, '..', 'scripts', 'nifty_chart_pattern_engine.py')
    
    spec = importlib.util.spec_from_file_location("nifty_engine", engine_path)
    if not spec:
         flash("Could not locate scripts/nifty_chart_pattern_engine.py", "danger")
         return redirect(url_for('admin.dashboard'))
         
    engine = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(engine)
    except Exception as e:
        flash(f"Error loading Pattern Engine module: {e}. Check server logs.", "danger")
        return redirect(url_for('admin.dashboard'))
        
    # Fetch data
    index_data = IndexData.query.filter(
        IndexData.symbol == symbol,
        IndexData.timestamp >= ts_start,
        IndexData.timestamp <= ts_end
    ).order_by(IndexData.timestamp).all()
    
    if len(index_data) < 10:
        if request.args:
            flash(f"Not enough candles ({len(index_data)}) for {symbol} on {date_str} between {start_time_str} and {end_time_str}.", "warning")
        return render_template('admin/live_pattern_match.html', 
                               today_curve=[], matches=[], today_stats={},
                               symbols=['NIFTY', 'BANKNIFTY'],
                               filters={'symbol': symbol, 'date': date_str, 'start_time': start_time_str, 'end_time': end_time_str, 'threshold': threshold})
                               
    # Transform to df
    import pandas as pd
    import numpy as np
    df = pd.DataFrame([{
        'timestamp': d.timestamp,
        'open': d.open or 0.0,
        'high': d.high or 0.0,
        'low': d.low or 0.0,
        'close': d.close or 0.0,
        'adx': d.adx,
        'supertrend_direction': d.supertrend_direction,
        'super_power': d.super_power
    } for d in index_data])
    
    # Run engine logic
    curve = engine.build_curve(df)
    features = engine.compute_features(df)
    ptype = engine.classify_pattern(df, curve)
    complexity = engine.curve_complexity(curve)
    
    # Load all patterns from DB matching symbol and ptype
    all_patterns = NiftyPattern.query.filter_by(symbol=symbol, pattern_type=ptype).all()
    
    results = []
    for p in all_patterns:
        pat_series = np.array(p.normalized_series)
        sim = engine.dtw_similarity(curve, pat_series)
        if sim >= threshold:
            results.append({
                'pattern': p,
                'similarity': round(sim, 2)
            })
            
    # Sort by similarity descending
    results.sort(key=lambda x: x['similarity'], reverse=True)
    
    # Top matches
    top_matches = results[:12]
    
    today_stats = {
        'pattern_type': ptype,
        'complexity': round(complexity, 2),
        'deviation_pct': features['deviation_pct'],
        'open_close_chg_pct': features['open_close_chg_pct'],
        'candles': len(df)
    }
    
    return render_template('admin/live_pattern_match.html',
                           today_curve=curve.tolist(),
                           today_stats=today_stats,
                           matches=top_matches,
                           symbols=['NIFTY', 'BANKNIFTY'],
                           filters={
                               'symbol': symbol, 
                               'date': date_str, 
                               'start_time': start_time_str, 
                               'end_time': end_time_str, 
                               'threshold': threshold
                           })
